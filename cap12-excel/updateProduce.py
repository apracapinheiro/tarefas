# Corrige os precos em uma planilha de venda de produtos.

import openpyxl


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# os tipos de produto e seus precos atualizados
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# percorre as linhas em um loop e atualiza os precos
for rowNum in range(2, sheet.max_row):  # pula a primeira linha
    produceName = sheet.cell(row=rowNum, column=1).value

    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')